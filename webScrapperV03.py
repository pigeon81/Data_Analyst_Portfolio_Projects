from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import *
import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\Lenovo\\Desktop\\scrp.xlsx')

sh1 = wb['Sheet1']
row = sh1.max_row

print("Funn has begun!!")

async def work(s, url):
    r = await s.get(url)

for i in range(1, 101):
    searchItem = sh1.cell(i,1).value
    
    driver = webdriver.Chrome( executable_path= "D:\\py\chromedriver.exe")
    url = "https://www.dnb.com/"
    driver.implicitly_wait(5)
    driver.get(url)
    driver.find_element_by_xpath('/html/body/div[1]/header/div[3]/div/div/div[2]/div[1]/button').click()
    driver.implicitly_wait(3)
    driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[2]/div/div/div[2]/div/div[1]/div/input').send_keys(searchItem)
    driver.implicitly_wait(2)
    #text = driver.find_element_by_xpath('//*[@id="company_search_results"]/ul/li/a[1]').text
    driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[2]/div/div/div[2]/div/div[1]/div/input').send_keys(Keys.ENTER)
    url2 = driver.current_url
    driver.quit
    print(url2)
    driver1 = webdriver.Chrome( executable_path= "D:\\py\chromedriver.exe")
    
    print("mauj!!")
    #driver1.implicitly_wait(4)

    try:

        driver1.get(url2) 
        driver1.implicitly_wait(3)
        #WebDriverWait(driver1, 30).until(expected_conditions.presence_of_all_elements_located(By.CLASS_NAME, 'primary_name' ))
        url3 = driver1.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[3]/div/div/div[3]/div/ul/li/div[1]/div[1]/a').get_attribute('href')
        driver1.implicitly_wait(3)
        #driver1.quit
        driver2 = webdriver.Chrome( executable_path= "D:\\py\chromedriver.exe")
        driver2.get(url3)
        driver1.close

        try: 
            outer = driver2.find_element_by_id('hero-company-link').get_attribute('href')        
            #website = outer.a.get_attribute('href')
            sh1.cell(row = i, column=2, value= outer)
        except:
            sh1.cell(row = i, column=2, value="Not found")
        try:
            location = driver2.find_element_by_class_name('company_region').text
            sh1.cell(row = i, column=3, value= location)
        except:
            sh1.cell(row = i, column=3, value="Not found")
        try : 
            
            contactNum= driver2.find_element_by_class_name('profile-phone-element').text
            sh1.cell(row = i, column=4, value = contactNum)
        except:
            sh1.cell(row = i, column=4, value="Not found")
        #driver.implicitly_wait(2)
        driver2.quit() 

    except:
        sh1.cell(row = i, column=2, value="Not found")
        sh1.cell(row = i, column=3, value="Not found")
        sh1.cell(row = i, column=4, value="Not found")


wb.save("C:\\Users\\Lenovo\\Desktop\\Report24.xlsx")    


